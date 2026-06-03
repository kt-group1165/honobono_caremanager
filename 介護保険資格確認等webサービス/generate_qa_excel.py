"""居宅介護支援事業所 ケアマネ目線 Q&A エクセル生成
   介護保険資格確認等WEBサービス + ケアプランデータ連携"""
import os
import sys
sys.stdout.reconfigure(encoding='utf-8')

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "ケアマネQ&A"

# ===== カラー =====
NAVY = "06365A"
TEAL = "028090"
WARM = "FFF3D6"
LIGHT = "F7F5F0"
CREAM = "FEF6EE"
ACCENT = "E86A0C"
WHITE = "FFFFFF"
GRAY_TEXT = "6B7280"

# ===== Q&A データ =====
# (カテゴリ, 質問, 回答, 補足/出典)
qa_data = [
    # ----- 制度の概要 -----
    ("制度の概要", "そもそも 介護保険資格確認等WEBサービス って何?",
     "国保中央会が運営する、介護事業所向けの Web API システムです。利用者の介護保険情報 (認定区分・有効期限・負担割合・住宅改修費利用状況など) を、自社の介護ソフトから電子的に取得できます。さらにケアプランや利用票・提供票を、サービス事業所と電子的にやり取りできるようになります。",
     "国保中央会・介護情報基盤等コンタクトセンター"),

    ("制度の概要", "いつから使えるの?",
     "令和8年 (2026年) 4月から暫定版がスタートします。介護情報の API と資格確認 API がまず使えるようになります。ケアプランデータ連携の本格機能 (事業所間連携・交付用登録) は令和8年度下期 (2026年10月頃) からです。",
     "段階的開始のため、当初は機能限定"),

    ("制度の概要", "国保連の既存ケアプラン連携システムとは別物?",
     "別物です。既存の「ケアプランデータ連携システム」(専用ソフトに CSV をアップロード) は引き続き使えますが、新しい WEBサービスは介護ソフトから直接 API 呼び出しで連携できる仕組みです。介護ソフトベンダーが対応すれば、ソフト内で全て完結します。",
     "既存システムとの併存期間あり"),

    ("制度の概要", "義務化されるの?使わないとダメ?",
     "現時点で利用は任意です。ただし、被保険者証の電子化・マイナ保険証への一本化の流れがあるため、将来的には事実上の標準になる可能性が高いです。早めの準備を推奨します。",
     "厚労省は普及促進方針"),

    # ----- マイナンバーカード読み取り -----
    ("マイナ読み取り", "利用者のマイナンバーカードはどう使うの?",
     "スマートフォン (NFC 対応) または PC のカードリーダーで、利用者のマイナンバーカードを読み取ります。読み取ると、その利用者の介護保険情報をオンラインで取得する「同意」が得られた状態になり、ケアマネが介護ソフトから情報を照会できるようになります。",
     "「マイナ資格確認アプリ」を使用"),

    ("マイナ読み取り", "毎月、毎回マイナンバーカードを読み取りに行くの?",
     "現時点での情報では、初回の同意取得 (マイナカード読み取り) 後は、被保険者番号で API 照会するだけで認定更新・負担割合変更等が自動反映される設計です。ただし、同意の有効期間・認定更新時の再取得要否・事業所変更時の取り扱い等の運用詳細は、令和8年4月の正式版で確定予定です。",
     "現時点情報 / 正式版で確定"),

    ("マイナ読み取り", "利用者がマイナンバーカードを持っていない場合は?",
     "現時点での情報では、代替手段として、被保険者番号 + 利用者基本情報 (氏名・生年月日等) による「資格確認用照会」も可能になる見込みです。ただし運用上は、マイナンバーカードを取得済みの利用者を優先的に切り替えていく形が現実的です。詳細は令和8年4月の正式版で確定予定。",
     "現時点情報 / 正式版で確定"),

    ("マイナ読み取り", "ケアマネが利用者の家でスマホで読み取ればOK?",
     "はい。ケアマネが訪問時にスマホで読み取れば、その場で同意取得が完了します。利用者にカードを持って事業所に来てもらう必要はありません。事業所内では PC + カードリーダーでも可能です。",
     "訪問業務との親和性が高い"),

    ("マイナ読み取り", "認知症で本人の同意が取れない場合は?",
     "現時点での情報では、ご家族等の代理同意の枠組みが用意される見込みです。詳細は令和8年4月の正式版で確定予定。現行の代行申請と同様の取り扱いになると想定されます。",
     "現時点情報 / 正式版で確定"),

    # ----- 業務上のメリット -----
    ("業務メリット", "結局、ケアマネにとって何が一番嬉しいの?",
     "1) 被保険者証・負担割合証のコピーを毎年もらわなくて良くなる、2) 認定有効期限を介護ソフトが自動アラートしてくれる、3) ケアプランや利用票・提供票を FAX しなくてよくなる (サービス事業所と電子連携)、の3点が最大効果です。",
     "月末月初の事務負担が大幅減"),

    ("業務メリット", "FAX 業務はどれくらい減る?",
     "ケアプラン連携が稼働すれば、サービス事業所への利用票・提供票・計画書配布 が電子化されます。月初の FAX 30〜50枚 / 一事業所あたりが想定される業務量とすると、その大半が削減対象になります。受け取り側 (訪問介護・福祉用具) も同様にメリットがあります。",
     "事業所間連携機能 (2026年10月〜)"),

    ("業務メリット", "他事業所のケアプラン履歴は見られる?",
     "「交付用登録機能」により、過去にその利用者に対して交付されたケアプランが介護情報基盤に保管されます。利用者が他の居宅事業所から移ってきた場合、過去のケアプラン情報を参照できる予定です (利用者の同意の範囲内)。",
     "引継ぎがスムーズに"),

    ("業務メリット", "住宅改修費の残額もわかるの?",
     "はい。住宅改修費 (上限20万円) の利用累計・残額や、福祉用具購入費 (年間10万円) の利用状況が、API でリアルタイム取得できます。これまで市区町村に電話確認していた業務が不要になります。",
     "計画立案の精度向上"),

    ("業務メリット", "認定更新の漏れチェックは?",
     "認定有効期限が API で取得できるため、介護ソフト側で「期限切れ 30日前アラート」などの機能が実装されます。期限切れによる算定エラー・返戻を未然に防げます。",
     "ベンダー実装依存"),

    # ----- 準備・コスト -----
    ("準備・コスト", "使うにはまず何が必要?",
     "(1) クライアント証明書の取得 (介護DX証明書 または 介護保険証明書)、(2) ユーザアカウントの払い出し (事業所ユーザ → 管理者ユーザ → 一般ユーザ の3階層)、(3) マイナ資格確認アプリの設定 (PC カードリーダー or スマホ NFC)、(4) 対応介護ソフトの準備、の4点です。",
     "申請窓口: 国保中央会"),

    ("準備・コスト", "介護DX証明書って何?既存の証明書と違う?",
     "介護DX証明書は令和7年10月から発行開始された新しいクライアント証明書です。既存の「介護保険証明書」(電子請求用) も当面は併用可能です。新規取得なら介護DX証明書が推奨されます。",
     "発行は国保連経由"),

    ("準備・コスト", "費用はどれくらいかかる?",
     "現時点で公表されているのは「証明書発行手数料」のみで、月額利用料は明示されていません。介護ソフト側の API 対応費用 (バージョンアップ費用) は各ベンダー次第です。詳細はベンダーへの確認が必要です。",
     "ベンダー個別問い合わせ"),

    ("準備・コスト", "今使っている介護ソフトで使える?",
     "各介護ソフトベンダーが対応 API の実装を進めています。お使いのベンダーに「介護保険資格確認等WEBサービス対応予定」を確認してください。対応していない場合は、令和8年度中の対応が見込まれます。",
     "ベンダー実装ロードマップ要確認"),

    ("準備・コスト", "PC環境の要件は?",
     "Windows PC + IC カードリーダー (公的個人認証対応) + ブラウザ環境 が基本構成です。スマートフォン (Android/iPhone NFC 対応) でも代用可能。事業所のネット環境は通常のブロードバンドで十分です。",
     "特別な専用回線は不要"),

    # ----- セキュリティ・同意 -----
    ("セキュリティ", "利用者の情報が漏れない?",
     "クライアント証明書 + ユーザ ID + パスワード + マイナンバーカード読み取りによる利用者同意 の多層認証です。同意の範囲外の情報は取得できません。また、ケアプラン連携は暗号化通信で送受信されます。",
     "国保連の厳格な認証基盤"),

    ("セキュリティ", "利用者が同意を撤回したらどうなる?",
     "利用者は同意を撤回でき、撤回後は API 経由での情報取得ができなくなります。これまで取得した情報の削除義務などの詳細は、正式仕様で確定予定です。",
     "同意管理は重要業務"),

    ("セキュリティ", "ケアマネ以外のスタッフも操作できる?",
     "事業所ユーザの管理者が「管理者ユーザ」「一般ユーザ」を発行できます。アクセスログも管理者が確認できます。役割に応じた権限管理が可能です。",
     "3階層のユーザ管理"),

    ("セキュリティ", "情報を見せていい範囲は?",
     "サービス事業所間の連携では、ケアプラン上で連携が必要な事業所 (担当者会議に出ているような事業所) にのみ情報が共有されます。連携範囲は利用者の同意に基づきます。",
     "ケアプラン上の連携先のみ"),

    # ----- 運用上の注意 -----
    ("運用注意", "API が止まったら業務が止まる?",
     "国保連側のメンテナンスや障害時には、一時的に照会できなくなります。重要な照会は事前に介護ソフトに保存される設計が一般的なので、業務継続には影響しない見込みです。",
     "業務継続性は配慮設計"),

    ("運用注意", "従来の紙の被保険者証は廃止?",
     "現時点では紙の被保険者証も併存します。マイナ保険証への一本化が進む中で、将来的には電子化が主流となる見込みです。当面は紙とデジタルの併用期間です。",
     "段階的移行"),

    ("運用注意", "サービス事業所がこの仕組みに未対応なら?",
     "従来通り FAX・郵送・手渡しで継続可能です。連携できる事業所から順次切り替える形が現実的です。福祉用具・訪問介護等の主要連携先から導入を進めると効果が高いです。",
     "段階導入を推奨"),

    ("運用注意", "ケアプランをアップロードするタイミングは?",
     "従来の「交付タイミング」(担当者会議後の確定版交付) と同じです。電子的に各サービス事業所へ送信され、サービス事業所側も介護ソフトで受信・確認できます。",
     "業務フローは変えなくて良い"),

    ("運用注意", "電話・口頭での連絡は不要になる?",
     "緊急時・状態変化時の電話連絡は引き続き重要です。電子連携は「定例的な書類のやり取り」を効率化するものであり、利用者対応のコミュニケーションは従来通り対人で行います。",
     "電子化は事務効率化が目的"),

    # ----- KTグループからの案内 -----
    ("KTグループ", "KTグループでは対応する予定?",
     "はい。KTグループ (居宅介護支援・訪問介護・福祉用具貸与) では、各事業所間の電子連携を進める予定です。同グループ内事業所との連携はスムーズに、グループ外事業所とも順次対応していきます。",
     "kt-group.co.jp"),

    ("KTグループ", "相談したい場合は?",
     "本制度の活用、対応介護ソフトの選定、連携先事業所との調整など、お気軽にご相談ください。KTグループ内の事業所であれば、共通の介護ソフトを使用しているケースが多く、導入支援も可能です。",
     "TEL 043-XXX-XXXX"),
]

# ===== レイアウト =====
# ヘッダー (タイトル + 説明)
ws.merge_cells("A1:D1")
c = ws["A1"]
c.value = "介護保険資格確認等WEBサービス  ケアマネ向け Q&A"
c.font = Font(name="游ゴシック", size=18, bold=True, color=WHITE)
c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
c.fill = PatternFill("solid", fgColor=NAVY)
ws.row_dimensions[1].height = 38

ws.merge_cells("A2:D2")
c = ws["A2"]
c.value = "令和8年4月運用開始 / 居宅介護支援事業所のケアマネ目線で「よくある疑問」を Q&A 形式でまとめました"
c.font = Font(name="游ゴシック", size=10, color=WHITE)
c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
c.fill = PatternFill("solid", fgColor=TEAL)
ws.row_dimensions[2].height = 22

# 空白行
ws.row_dimensions[3].height = 8

# テーブルヘッダ
headers = ["カテゴリ", "Q (質問)", "A (回答)", "補足 / 出典"]
header_row = 4
for i, h in enumerate(headers, start=1):
    cell = ws.cell(row=header_row, column=i, value=h)
    cell.font = Font(name="游ゴシック", size=11, bold=True, color=WHITE)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.fill = PatternFill("solid", fgColor=NAVY)
    cell.border = Border(
        left=Side(style="thin", color="FFFFFF"),
        right=Side(style="thin", color="FFFFFF"),
        top=Side(style="medium", color=NAVY),
        bottom=Side(style="medium", color=NAVY),
    )
ws.row_dimensions[header_row].height = 32

# カテゴリ色マップ
cat_colors = {
    "制度の概要": "DBEAFE",
    "マイナ読み取り": "FEF3C7",
    "業務メリット": "DCFCE7",
    "準備・コスト": "FCE7F3",
    "セキュリティ": "FEE2E2",
    "運用注意": "E0E7FF",
    "KTグループ": "FFF3D6",
}

thin_border = Border(
    left=Side(style="thin", color="D1D5DB"),
    right=Side(style="thin", color="D1D5DB"),
    top=Side(style="thin", color="E5E7EB"),
    bottom=Side(style="thin", color="E5E7EB"),
)

# データ行
prev_cat = None
for idx, (cat, q, a, ref) in enumerate(qa_data, start=1):
    row = header_row + idx
    # カテゴリ
    cc = ws.cell(row=row, column=1, value=cat if cat != prev_cat else "")
    cc.font = Font(name="游ゴシック", size=10, bold=True, color=NAVY)
    cc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cc.fill = PatternFill("solid", fgColor=cat_colors.get(cat, "F3F4F6"))
    cc.border = thin_border
    # Q
    qc = ws.cell(row=row, column=2, value=f"Q{idx:02d}. {q}")
    qc.font = Font(name="游ゴシック", size=10.5, bold=True, color="111827")
    qc.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
    qc.fill = PatternFill("solid", fgColor=LIGHT)
    qc.border = thin_border
    # A
    ac = ws.cell(row=row, column=3, value=a)
    ac.font = Font(name="游ゴシック", size=10, color="1F2937")
    ac.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
    ac.fill = PatternFill("solid", fgColor=WHITE)
    ac.border = thin_border
    # 補足
    rc = ws.cell(row=row, column=4, value=ref)
    rc.font = Font(name="游ゴシック", size=9, italic=True, color=GRAY_TEXT)
    rc.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
    rc.fill = PatternFill("solid", fgColor=CREAM)
    rc.border = thin_border
    # 行高 (内容長に応じて)
    a_lines = max(1, len(a) // 38 + a.count("\n"))
    q_lines = max(1, len(q) // 30)
    ws.row_dimensions[row].height = max(48, max(a_lines, q_lines) * 16 + 14)

    prev_cat = cat

# フッター
foot_row = header_row + len(qa_data) + 2
ws.merge_cells(start_row=foot_row, start_column=1, end_row=foot_row, end_column=4)
c = ws.cell(row=foot_row, column=1)
c.value = "※ 本資料は現時点の情報 (国保中央会公表の暫定版仕様) に基づき作成。同意の有効期間・再取得要否・代理同意の取り扱い等の運用詳細は、令和8年4月の正式版で確定予定です。お問い合わせ: KTグループ / 株式会社ケイ・ティ・サービス (kt-group.co.jp)"
c.font = Font(name="游ゴシック", size=9, color=GRAY_TEXT, italic=True)
c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
c.fill = PatternFill("solid", fgColor=LIGHT)
ws.row_dimensions[foot_row].height = 30

# カラム幅
ws.column_dimensions["A"].width = 16   # カテゴリ
ws.column_dimensions["B"].width = 42   # Q
ws.column_dimensions["C"].width = 70   # A
ws.column_dimensions["D"].width = 26   # 補足

# 印刷設定 (A4 横)
ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
ws.page_setup.paperSize = ws.PAPERSIZE_A4
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 0
ws.sheet_properties.pageSetUpPr.fitToPage = True
ws.page_margins.left = 0.4
ws.page_margins.right = 0.4
ws.page_margins.top = 0.5
ws.page_margins.bottom = 0.5
ws.print_options.horizontalCentered = True
# 各ページにヘッダ行を繰り返し
ws.print_title_rows = f"1:{header_row}"

# ウィンドウ枠の固定 (ヘッダ行)
ws.freeze_panes = f"A{header_row + 1}"

# ===== シート2: カテゴリ別サマリ =====
ws2 = wb.create_sheet("カテゴリ別サマリ")

ws2.merge_cells("A1:C1")
c = ws2["A1"]
c.value = "カテゴリ別 質問数"
c.font = Font(name="游ゴシック", size=16, bold=True, color=WHITE)
c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
c.fill = PatternFill("solid", fgColor=NAVY)
ws2.row_dimensions[1].height = 32

# サマリ集計
from collections import Counter
cat_counts = Counter(c for c, _, _, _ in qa_data)

ws2.cell(row=3, column=1, value="カテゴリ").font = Font(name="游ゴシック", size=11, bold=True, color=WHITE)
ws2.cell(row=3, column=2, value="質問数").font = Font(name="游ゴシック", size=11, bold=True, color=WHITE)
ws2.cell(row=3, column=3, value="代表的なテーマ").font = Font(name="游ゴシック", size=11, bold=True, color=WHITE)
for col in range(1, 4):
    cell = ws2.cell(row=3, column=col)
    cell.fill = PatternFill("solid", fgColor=NAVY)
    cell.alignment = Alignment(horizontal="center", vertical="center")
ws2.row_dimensions[3].height = 26

themes = {
    "制度の概要": "サービスの全体像・開始時期・既存仕組みとの違い",
    "マイナ読み取り": "カードの使い方・読み取り頻度・代替手段",
    "業務メリット": "FAX削減・ケアプラン履歴参照・住宅改修費確認",
    "準備・コスト": "証明書・アカウント・対応介護ソフト・費用",
    "セキュリティ": "情報漏えい対策・同意管理・アクセス権限",
    "運用注意": "API障害時・電話連絡・サービス事業所未対応",
    "KTグループ": "グループ内連携・相談窓口",
}

cat_order = ["制度の概要", "マイナ読み取り", "業務メリット", "準備・コスト", "セキュリティ", "運用注意", "KTグループ"]
for i, cat in enumerate(cat_order, start=4):
    cell = ws2.cell(row=i, column=1, value=cat)
    cell.font = Font(name="游ゴシック", size=11, bold=True, color=NAVY)
    cell.fill = PatternFill("solid", fgColor=cat_colors.get(cat, "F3F4F6"))
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border
    cell = ws2.cell(row=i, column=2, value=cat_counts[cat])
    cell.font = Font(name="游ゴシック", size=14, bold=True, color=ACCENT)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border
    cell = ws2.cell(row=i, column=3, value=themes[cat])
    cell.font = Font(name="游ゴシック", size=10, color="1F2937")
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
    cell.border = thin_border
    ws2.row_dimensions[i].height = 32

# 合計
total_row = 4 + len(cat_order)
cell = ws2.cell(row=total_row, column=1, value="合計")
cell.font = Font(name="游ゴシック", size=11, bold=True, color=WHITE)
cell.fill = PatternFill("solid", fgColor=TEAL)
cell.alignment = Alignment(horizontal="center", vertical="center")
cell = ws2.cell(row=total_row, column=2, value=sum(cat_counts.values()))
cell.font = Font(name="游ゴシック", size=14, bold=True, color=WHITE)
cell.fill = PatternFill("solid", fgColor=TEAL)
cell.alignment = Alignment(horizontal="center", vertical="center")
cell = ws2.cell(row=total_row, column=3, value=f"Q01 〜 Q{sum(cat_counts.values()):02d}")
cell.font = Font(name="游ゴシック", size=10, bold=True, color=WHITE)
cell.fill = PatternFill("solid", fgColor=TEAL)
cell.alignment = Alignment(horizontal="center", vertical="center")
ws2.row_dimensions[total_row].height = 30

ws2.column_dimensions["A"].width = 18
ws2.column_dimensions["B"].width = 12
ws2.column_dimensions["C"].width = 60

# 保存
OUT = os.path.join(os.path.dirname(__file__), 'qa_caremanager.xlsx')
try:
    wb.save(OUT)
except PermissionError:
    OUT = os.path.join(os.path.dirname(__file__), 'qa_caremanager_v2.xlsx')
    wb.save(OUT)

print(f"生成完了: {OUT}")
print(f"Q数: {len(qa_data)} / カテゴリ数: {len(cat_counts)}")
