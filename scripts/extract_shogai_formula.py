"""障害サービス報酬算定構造.xlsx から処遇改善加算等の率を抽出し、
   既存 service_codes_shogai.csv に formula 列を埋めて出力する。

   - 各シート (33 サービス) で「＋所定単位×N／1,000」or「＋所定単位×N/1,000」表記を検出
   - 加算名 (Ⅰ/Ⅱ/Ⅲ/Ⅳ/Ⅴ(1)〜(14)) と対応付け
   - service_codes_shogai.csv の該当行 (= 「○処遇改善加算Ⅰ」 等) に formula JSON を埋める
"""

import csv
import json
import re
import sys
import openpyxl

sys.stdout.reconfigure(encoding="utf-8")

XLSX = "C:/Users/domen-PC/Downloads/介護システム統合/docs/介護ソフト関連/障害/障害サービス_報酬算定構造.xlsx"
CSV_IN = "service_codes_shogai.csv"
CSV_OUT = "service_codes_shogai.csv"

# シート名 → service_category prefix (extract_service_codes.mjs と整合)
SHEET_TO_CATEGORY = {
    "１．居宅介護": "11",
    "２．重度訪問": "12",
    "３．同行援護 ": "15",
    "４．行動援護 ": "13",
    "５．重度包括": "14",
    "６．療養介護": "21",
    "７．生活介護": "22",
    "８．短期入所": "24",
    "９．施設入所": "32",
    "１０．共同生活援助": "33",
    "１１．自立生活援助": "35",
    "１２．機能訓練": "41",
    "１３．生活訓練": "42",
    "１４．宿泊型自立訓練": "34",  # 一部別カテゴリ
    "１５．就労移行 ": "43",
    "１６．就労移行(養成) ": "44",
    "１６．就労移行(養成)　": "44",
    "１６．就労移行（養成） ": "44",
    "１７．就労Ａ型 ": "45",
    "１８．就労Ｂ型 ": "46",
    "１９．就労定着支援 ": "47",
    "２０．就労選択": "48",
    "２１．計画相談支援": "52",
    "２２．障害児相談支援": "55",
    "２３．地域移行支援": "53",
    "２４．地域定着支援": "54",
    "２５．児童発達支援 ": "61",
    "２６．児童発達支援  (難聴)": "61",
    "２７．児童発達支援  (重症児)": "61",
    "２８．旧医療型児童発達支援": "62",
    "２９．放課後等デイサービス": "63",
    "３０．居宅訪問型児童発達支援": "65",
    "３１．保育所等訪問支援": "64",
    "３２．福祉型障害児入所": "71",
    "３３．医療型障害児入所": "72",
}

# 加算名キーワード → 派生インデックス
RANK_PATTERNS = [
    ("Ⅰ", r"処遇改善加算\s*[（(]?Ⅰ[）)]?\s*$"),
    ("Ⅱ", r"処遇改善加算\s*[（(]?Ⅱ[）)]?\s*$"),
    ("Ⅲ", r"処遇改善加算\s*[（(]?Ⅲ[）)]?\s*$"),
    ("Ⅳ", r"処遇改善加算\s*[（(]?Ⅳ[）)]?\s*$"),
]


def parse_sheet(ws):
    """シート内の福祉・介護職員等処遇改善加算の Ⅰ〜Ⅳ 率を抽出"""
    rates = {}  # rank("Ⅰ"〜"Ⅳ") → (num, den)
    last_rank = None
    is_in_shogu_section = False
    for r in range(1, ws.max_row + 1):
        for c in range(1, min(ws.max_column + 1, 10)):
            v = ws.cell(row=r, column=c).value
            if not v:
                continue
            sv = str(v)
            # セクション開始判定
            if "福祉・介護職員等処遇改善加算" == sv.strip() or "福祉・介護職員等処遇改善加算" in sv and "加算" in sv and "Ⅰ" not in sv and "Ⅱ" not in sv and "Ⅴ" not in sv:
                # ヘッダー行を許容
                is_in_shogu_section = True
            if "福祉・介護職員処遇改善加算" in sv and "等" not in sv:
                # 旧加算セクション開始 → 等処遇改善セクション終了
                is_in_shogu_section = False
            # 各 rank 名
            if is_in_shogu_section:
                for rank, pat in RANK_PATTERNS:
                    if re.search(pat, sv):
                        # 既に rank の率が見つかってる場合は skip (Ⅴ(1) などは別)
                        if rank not in rates:
                            last_rank = rank
                            break
            # 率表記 (カンマ区切り対応: "1,000" を考慮)
            m = re.search(
                r"所定単位\s*[×x]\s*(\d{1,4}(?:[,，]\d{3})*)\s*[／/]\s*(\d{1,3}(?:[,，]\d{3})*)",
                sv,
            )
            if m and last_rank and is_in_shogu_section:
                if last_rank not in rates:
                    num = int(m.group(1).replace(",", "").replace("，", ""))
                    den = int(m.group(2).replace(",", "").replace("，", ""))
                    rates[last_rank] = (num, den)
                last_rank = None  # 1 rank 1 rate
    return rates


def main():
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    # シート別に率を抽出
    category_rates = {}  # category → {rank → (num, den)}
    for sheet_name in wb.sheetnames:
        cat = SHEET_TO_CATEGORY.get(sheet_name)
        if not cat:
            continue
        rates = parse_sheet(wb[sheet_name])
        if rates:
            # 同じ category に複数 sheet があれば上書き (児発系は同 cat 61 で重複可、最初を採用)
            if cat not in category_rates:
                category_rates[cat] = rates
            else:
                # マージ (欠けてる rank だけ追加)
                for k, v in rates.items():
                    category_rates[cat].setdefault(k, v)
            print(f"  [{cat}] {sheet_name}: {rates}")

    print()
    print(f"=== 抽出済 {len(category_rates)} カテゴリ ===")

    # CSV を読んで formula 列を埋める
    with open(CSV_IN, encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        rows = list(reader)
        headers = reader.fieldnames

    updated = 0
    for r in rows:
        if r["calculation_type"] != "加算":
            continue
        if "処遇改善加算" not in r["service_name"]:
            continue
        cat = r["service_category"]
        if cat not in category_rates:
            continue
        # サービス名の末尾から Ⅰ/Ⅱ/Ⅲ/Ⅳ を判定 (例: 居介処遇改善加算Ⅰ)
        name = r["service_name"]
        rank = None
        for rk in ("Ⅰ", "Ⅱ", "Ⅲ", "Ⅳ"):
            if name.endswith(rk):
                rank = rk
                break
        if not rank:
            continue
        rate = category_rates[cat].get(rank)
        if not rate:
            continue
        num, den = rate
        r["formula"] = json.dumps({
            "type": "monthly_aggregate",
            "service_category": cat,
            "numerator": num,
            "denominator": den,
            "label": f"所定単位×{num}/{den}",
            "rounding": "round",
        }, ensure_ascii=False)
        updated += 1

    # 書き戻し
    with open(CSV_OUT, "w", encoding="utf-8-sig", newline="") as f:
        # BOM 付き UTF-8 で出力
        w = csv.DictWriter(f, fieldnames=headers, quoting=csv.QUOTE_MINIMAL)
        w.writeheader()
        w.writerows(rows)

    print(f"\n✓ formula を {updated} 件埋めて {CSV_OUT} に保存しました")


if __name__ == "__main__":
    main()
